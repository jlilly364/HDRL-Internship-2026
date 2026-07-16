import openpyxl
from openpyxl.styles import PatternFill
import ast

def highlight_new_sheet(file_path, output_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Define colors
    blue_fill = PatternFill(start_color="FFBDD7EE", end_color="FFBDD7EE", fill_type="solid") # Light Blue
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # Light Red

    # Change these letters to exactly where your data is in the NEW spreadsheet!
    col_name = 'E' # The column with the Name
    col_list = 'F' # The column with the List

    # Convert letters to 0-based index for openpyxl
    idx_name = openpyxl.utils.column_index_from_string(col_name) - 1
    idx_list = openpyxl.utils.column_index_from_string(col_list) - 1

    for row in ws.iter_rows(min_row=2, max_col=max(ws.max_column, idx_list + 1)):
        
        val_name = row[idx_name].value
        val_list = row[idx_list].value

        if not isinstance(val_list, str):
            continue

        if val_list.strip().startswith('['):
            try:
                list_data = ast.literal_eval(val_list.strip())
                val_name_clean = str(val_name).strip() if val_name else ""
                
                is_empty_list = (list_data == [''])
                name_is_not_first = (len(list_data) > 0 and val_name_clean != str(list_data[0]).strip())

                if is_empty_list:
                    # First highlight condition -> Light Blue
                    for cell in row:
                        cell.fill = blue_fill
                elif name_is_not_first:
                    # Second highlight condition -> Light Red
                    for cell in row:
                        cell.fill = red_fill

            except (ValueError, SyntaxError):
                pass

    wb.save(output_path)
    print("Done! Highlights applied to the new spreadsheet.")

# Run it on your new spreadsheet!
# Just replace these filenames with your actual files
highlight_new_sheet('copy_all_helioData_spase_names.xlsx', 'new_spreadsheet_highlighted.xlsx')